
import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

path = r'C:\Users\Laptop\Downloads\【2026】ITAC_P2P_02 采购收货入应付暂估会计凭证.xlsx'

def get_merged_value(sheet, cell):
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return sheet.cell(merged_range.min_row, merged_range.min_col).value
    return cell.value

try:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.worksheets[0]
    print(f'Sheet Name: {sheet.title}')
    
    print('--- Data Grid (Rows 1-20, Cols A-J) ---')
    for r in range(1, 21):
        row_values = []
        for c in range(1, 11):
            cell = sheet.cell(row=r, column=c)
            val = get_merged_value(sheet, cell)
            row_values.append(str(val) if val is not None else "")
        print(f"Row {r:2}: {' | '.join(row_values)}")

    print('\n--- Colors and Borders (Checking Row 6 and 15) ---')
    for r in [6, 15]:
        for c in range(1, 11):
            cell = sheet.cell(row=r, column=c)
            fill = cell.fill
            if fill and fill.start_color and fill.start_color.index != '00000000':
                 color = fill.start_color.rgb if fill.start_color.type == 'rgb' else f"Indexed {fill.start_color.index}"
                 print(f"Cell {cell.coordinate}: Color={color}")

except Exception as e:
    print(f"Error: {e}")
