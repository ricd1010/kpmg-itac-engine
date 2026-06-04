import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class ReportGenerator:
    def __init__(self, output_dir):
        self.output_path = os.path.join(output_dir, "WorkingPaper_Final.xlsx")

    def generate(self, ranked_scenarios, di_results, audit_context=None):
        if audit_context is None:
            audit_context = {}

        wb = Workbook()
        
        # 1. Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "测试汇总"
        
        # Style Definitions
        kpmg_blue_fill = PatternFill(start_color="00338D", end_color="00338D", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Title
        ws_summary.merge_cells("A1:E1")
        ws_summary["A1"] = "自动化控制 (ITAC) 测试汇总表"
        ws_summary["A1"].font = Font(size=16, bold=True, color="00338D")
        ws_summary["A1"].alignment = align_center

        # Headers
        table_headers = ["控制编号", "控制描述 / 审计场景", "风险等级", "涉及交易规模", "审计结论"]
        ws_summary.append(table_headers)
        for cell in ws_summary[ws_summary.max_row]:
            cell.font = header_font
            cell.fill = kpmg_blue_fill
            cell.alignment = align_center
            cell.border = thin_border

        # Data
        di_map = {}
        for res in di_results:
            name = res['scenario']
            if name not in di_map: di_map[name] = []
            di_map[name].append(res)

        for res in ranked_scenarios:
            name = res['name']
            samples = di_map.get(name, [])
            conclusion = f"测试满意 ({len(samples)}个样本)" if samples else "未执行测试"
            row = ["ITAC_CTRL", name, "High", f"{res['total_value']:,.2f}", conclusion]
            ws_summary.append(row)
            for cell in ws_summary[ws_summary.max_row]:
                cell.border = thin_border
        
        ws_summary.column_dimensions['A'].width = 15
        ws_summary.column_dimensions['B'].width = 50
        ws_summary.column_dimensions['D'].width = 20
        ws_summary.column_dimensions['E'].width = 25

        # 2. Detailed D&I Sheets (Based on Standard Template)
        for scenario_name, results in di_map.items():
            # Create a safe sheet name
            safe_title = "".join([c for c in scenario_name if c.isalnum() or c==' '])[:31].strip()
            ws = wb.create_sheet(title=safe_title)
            
            # Formatting as per template
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 30
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 60

            # Row 2: Header
            ws.merge_cells("D2:H2")
            ws["D2"] = "设计和执行(D&I)"
            ws["D2"].font = Font(size=14, bold=True)
            ws["D2"].alignment = Alignment(horizontal="center")

            # Row 3: Sub-header
            ws.merge_cells("D3:H3")
            ws["D3"] = "了解流程控制活动 / Understand the process control activities"
            ws["D3"].font = Font(bold=True)
            ws["D3"].alignment = Alignment(horizontal="center")

            # Row 6: Control Label
            ws["D6"] = "控制 / Control"
            ws["D6"].fill = kpmg_blue_fill
            ws["D6"].font = header_font
            ws["D6"].border = thin_border

            # Row 8: Control ID and Description
            ws["D8"] = "ITAC_GEN" # Placeholder ID
            ws["D8"].border = thin_border
            ws["E8"] = scenario_name
            ws["E8"].border = thin_border
            ws["E8"].alignment = Alignment(wrap_text=True)

            # Row 17: Table Headers
            cols = ["D", "E", "F", "G", "H"]
            headers = ["流程风险点 / PRP ID", "流程风险点 / PRP(s)", "重大错报风险 / RMM", "信息 / Information", "该流程控制活动如何应对流程风险点(PRP)"]
            for col, text in zip(cols, headers):
                ws[f"{col}17"] = text
                ws[f"{col}17"].fill = kpmg_blue_fill
                ws[f"{col}17"].font = Font(bold=True, color="FFFFFF", size=9)
                ws[f"{col}17"].alignment = align_center
                ws[f"{col}17"].border = thin_border

            # Row 18: TOD Narrative (Use first sample's D&I text)
            ws[f"D18"] = "PRP_01"
            ws[f"H18"] = results[0]['di_description']
            ws[f"H18"].alignment = Alignment(wrap_text=True, vertical="top")
            for col in cols: ws[f"{col}18"].border = thin_border

            # Row 20 & 22: Nature and Type
            ws["D20"] = "性质 / Nature"
            ws["D20"].fill = kpmg_blue_fill
            ws["D20"].font = header_font
            ws["E20"] = "自动化【AUTOMATED】"
            ws["D22"] = "类型 / Type"
            ws["D22"].fill = kpmg_blue_fill
            ws["D22"].font = header_font
            ws["E22"] = "预防性【PREVENTIVE】"
            ws["D20"].border = ws["E20"].border = ws["D22"].border = ws["E22"].border = thin_border

            # TOE Section (Sample List) below the D&I grid
            start_row = 26
            ws.cell(row=start_row, column=4, value="二、测试样本明细 (TOE)").font = Font(bold=True)
            headers_toe = ["凭证号", "科目", "描述", "金额", "日期"]
            for i, h in enumerate(headers_toe):
                cell = ws.cell(row=start_row+1, column=4+i, value=h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type="solid")
                cell.border = thin_border
            
            for idx, res in enumerate(results):
                row_idx = start_row + 2 + idx
                s = res['sample_table']
                ws.cell(row=row_idx, column=4, value=s['DOC_NUM']).border = thin_border
                ws.cell(row=row_idx, column=5, value=s['DEBIT_ACC']).border = thin_border
                ws.cell(row=row_idx, column=6, value=s['DEBIT_DESC']).border = thin_border
                ws.cell(row=row_idx, column=7, value=s['AMOUNT']).border = thin_border
                ws.cell(row=row_idx, column=8, value=s.get('DATE', 'N/A')).border = thin_border

        wb.save(self.output_path)
        return self.output_path
