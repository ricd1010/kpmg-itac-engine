import os
import openpyxl

path = '【2026】ITAC_P2P_02 采购收货入应付暂估会计凭证.xlsx'

def get_merged_value(sheet, row, col):
    cell = sheet.cell(row=row, column=col)
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return sheet.cell(merged_range.min_row, merged_range.min_col).value
    return cell.value

def get_cell_style(sheet, row, col):
    cell = sheet.cell(row=row, column=col)
    fill = cell.fill
    color = "None"
    if fill and fill.start_color:
        if fill.start_color.type == 'rgb':
            color = fill.start_color.rgb
        elif fill.start_color.type == 'indexed':
            color = f"Indexed {fill.start_color.index}"
    
    font = cell.font
    font_str = f"Bold={font.bold}, Color={font.color.rgb if font.color and font.color.type == 'rgb' else 'Default'}"
    
    return f"Color={color}, Font={font_str}"

try:
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"Sheet Names: {wb.sheetnames}")
    sheet = wb.worksheets[0]
    
    print("--- Detailed Grid (Rows 1-30) ---")
    for r in range(1, 31):
        row_vals = []
        for c in range(1, 15):
            val = get_merged_value(sheet, r, c)
            row_vals.append(str(val) if val is not None else "")
        print(f"R{r:2}: {' | '.join(row_vals)}")

    print("\n--- Header Styles Analysis ---")
    # Row 17 seems to be a major header row
    for c in range(3, 10): # Checking Col C to I
        val = get_merged_value(sheet, 17, c)
        if val:
            style = get_cell_style(sheet, 17, c)
            print(f"Row 17 Col {c} ('{val}'): {style}")

    # Check for D&I narrative or TOD/TOE
    print("\n--- Searching for Narrative ---")
    for r in range(1, 100):
        for c in range(1, 15):
            val = get_merged_value(sheet, r, c)
            if val and isinstance(val, str):
                if any(keyword in val for keyword in ["TOD", "TOE", "Test of Design", "Test of Effectiveness", "测试过程", "穿行测试"]):
                    print(f"Found narrative at R{r}C{c}: {val[:200]}...")

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
