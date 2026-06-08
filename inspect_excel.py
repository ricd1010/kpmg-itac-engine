import os
import openpyxl
import sys

# Set encoding for output
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

path = '【2026】ITAC_P2P_02 采购收货入应付暂估会计凭证.xlsx'

try:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.worksheets[0]
    print(f'Sheet Name: {sheet.title}')
    
    print('--- Top 15 rows ---')
    for row in sheet.iter_rows(max_row=15, values_only=True):
        print(row)
        
    print('--- Header Styles (Row 6) ---')
    # Let's check row 6 specifically as it might be the table header
    for cell in sheet[6]:
        if cell.value:
            color = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color.type == 'rgb' else "No Color"
            if cell.fill and cell.fill.start_color.type == 'indexed':
                 color = f"Indexed {cell.fill.start_color.index}"
            print(f"Cell {cell.column_letter}6: '{cell.value}' - Color: {color}")

    # Check for D&I narrative
    found_narrative = False
    for row in sheet.iter_rows(max_row=50, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str) and ('TOD' in cell or 'TOE' in cell or '穿行测试' in cell):
                found_narrative = True
                print(f"Found narrative-like content: {cell[:100]}...")
                break
        if found_narrative:
            break

except Exception as e:
    print(f"Error: {e}")
