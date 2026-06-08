import os
import openpyxl

path = "【2026】ITAC_P2P_02 采购收货入应付暂估会计凭证.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
print(f"Sheets: {wb.sheetnames}")

# Analyze the first sheet specifically
ws = wb[wb.sheetnames[0]]
print(f"\n--- 正在分析第一个分页: {wb.sheetnames[0]} ---")
for row in range(1, 25):
    row_data = []
    for col in range(1, 10):
        cell = ws.cell(row=row, column=col)
        val = cell.value
        row_data.append(f"[{val}]")
    print(f"Row {row:2}: {' '.join(row_data)}")
