import os
import re
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class ReportGenerator:
    """Generate V.A.S.T. SAP automated-voucher audit working papers."""

    KPMG_BLUE = "00338D"
    LIGHT_BLUE = "EAF3FF"
    LIGHT_TEAL = "EAF7F7"
    LIGHT_GREY = "F7F9FC"
    MID_GREY = "D8DDE6"

    def __init__(self, output_dir):
        self.output_dir = str(output_dir)
        self.output_path = os.path.join(self.output_dir, "WorkingPaper_Final.xlsx")

    @staticmethod
    def _text(value, default=""):
        if isinstance(value, (list, tuple, set)):
            value = next((item for item in value if str(item).strip()), "")
        elif isinstance(value, dict):
            value = next((item for item in value.values() if str(item).strip()), "")
        try:
            if pd.isna(value):
                return default
        except (TypeError, ValueError):
            return default
        text = str(value).strip()
        return default if not text or text.lower() in {"nan", "none", "null"} else text

    @classmethod
    def _num(cls, value):
        if isinstance(value, (list, tuple, set, dict)):
            return 0.0
        text = cls._text(value)
        if not text:
            return 0.0
        negative = text.startswith("(") and text.endswith(")") or text.endswith("-") or text.startswith("-")
        cleaned = text.replace(",", "").replace("CNY", "").replace("RMB", "").strip("()- ")
        try:
            number = float(cleaned)
        except ValueError:
            return 0.0
        return -abs(number) if negative else number

    @staticmethod
    def _styles():
        border = Border(
            left=Side(style="thin", color=ReportGenerator.MID_GREY),
            right=Side(style="thin", color=ReportGenerator.MID_GREY),
            top=Side(style="thin", color=ReportGenerator.MID_GREY),
            bottom=Side(style="thin", color=ReportGenerator.MID_GREY),
        )
        return {
            "border": border,
            "blue_fill": PatternFill(start_color=ReportGenerator.KPMG_BLUE, end_color=ReportGenerator.KPMG_BLUE, fill_type="solid"),
            "light_blue_fill": PatternFill(start_color=ReportGenerator.LIGHT_BLUE, end_color=ReportGenerator.LIGHT_BLUE, fill_type="solid"),
            "light_teal_fill": PatternFill(start_color=ReportGenerator.LIGHT_TEAL, end_color=ReportGenerator.LIGHT_TEAL, fill_type="solid"),
            "light_grey_fill": PatternFill(start_color=ReportGenerator.LIGHT_GREY, end_color=ReportGenerator.LIGHT_GREY, fill_type="solid"),
            "header_font": Font(bold=True, color="FFFFFF"),
            "title_font": Font(size=16, bold=True, color=ReportGenerator.KPMG_BLUE),
            "section_font": Font(size=12, bold=True, color=ReportGenerator.KPMG_BLUE),
            "bold_font": Font(bold=True),
            "wrap": Alignment(wrap_text=True, vertical="top"),
            "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        }

    @staticmethod
    def _safe_sheet_title(title, used=None, prefix=""):
        used = used if used is not None else set()
        raw = f"{prefix}{title}".strip() or "未命名"
        base = re.sub(r"[\[\]\:\*\?\/\\]", "_", raw)[:31].strip() or "未命名"
        candidate = base
        idx = 2
        while candidate in used:
            suffix = f"_{idx}"
            candidate = f"{base[:31 - len(suffix)]}{suffix}"
            idx += 1
        used.add(candidate)
        return candidate

    @staticmethod
    def _column_widths(ws, widths):
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    @classmethod
    def _write_table(cls, ws, start_row, start_col, headers, rows, styles, title=None):
        row = start_row
        if title:
            ws.cell(row=row, column=start_col, value=title).font = styles["section_font"]
            row += 1
        for offset, header in enumerate(headers):
            cell = ws.cell(row=row, column=start_col + offset, value=header)
            cell.fill = styles["blue_fill"]
            cell.font = styles["header_font"]
            cell.alignment = styles["center"]
            cell.border = styles["border"]
        if not rows:
            row += 1
            cell = ws.cell(row=row, column=start_col, value="暂无数据 / 待补充")
            cell.alignment = styles["wrap"]
            cell.border = styles["border"]
            return row
        for values in rows:
            row += 1
            for offset, value in enumerate(values):
                cell = ws.cell(row=row, column=start_col + offset, value=value)
                cell.alignment = styles["wrap"]
                cell.border = styles["border"]
                if isinstance(value, (int, float)) and offset != 0:
                    cell.number_format = "#,##0.00"
        return row

    def _load_session_table(self, name):
        path = os.path.join(self.output_dir, f"{name}.csv")
        if not os.path.exists(path):
            return pd.DataFrame()
        for encoding in ("utf-8-sig", "utf-8", "gbk"):
            try:
                return pd.read_csv(path, dtype=str, encoding=encoding)
            except Exception:
                continue
        return pd.DataFrame()

    @classmethod
    def _scenario_total(cls, scenario):
        if "total_value" in scenario:
            return cls._num(scenario.get("total_value"))
        total = 0.0
        for company in scenario.get("company_values", []) or []:
            total += cls._num(company.get("total_value"))
            for account in company.get("account_values", []) or []:
                total += cls._num(account.get("amount")) or cls._num(account.get("total_value"))
        return total

    @classmethod
    def _account_rows(cls, ranked_scenarios):
        rows = []
        for scenario in ranked_scenarios or []:
            scenario_name = cls._text(scenario.get("name"))
            for company in scenario.get("company_values", []) or []:
                company_code = cls._text(company.get("company_code"), "未指定公司")
                for account in company.get("account_values", []) or []:
                    amount = (
                        cls._num(account.get("combined_value"))
                        or cls._num(account.get("total_value"))
                        or cls._num(account.get("amount"))
                        or cls._num(account.get("value"))
                    )
                    if not amount:
                        continue
                    rows.append({
                        "scenario": scenario_name,
                        "company": company_code,
                        "account": cls._text(account.get("account")),
                        "description": cls._text(account.get("description"), "未知科目"),
                        "amount": amount,
                        "debit": cls._num(account.get("debit_value")),
                        "credit": cls._num(account.get("credit_value")),
                    })
        return rows

    @classmethod
    def _top_accounts(cls, ranked_scenarios, limit=10):
        totals = defaultdict(float)
        descriptions = {}
        for row in cls._account_rows(ranked_scenarios):
            key = (row["scenario"], row["account"])
            totals[key] += row["amount"]
            descriptions[key] = row["description"]
        ranked = sorted(totals.items(), key=lambda item: abs(item[1]), reverse=True)
        return [
            {
                "scenario": scenario,
                "account": account,
                "description": descriptions.get((scenario, account), ""),
                "amount": amount,
            }
            for (scenario, account), amount in ranked[:limit]
        ]

    @classmethod
    def _scenario_account_details(cls, scenario):
        rows = []
        for detail in scenario.get("account_details", []) or []:
            rows.append({
                "direction": cls._text(detail.get("direction")),
                "account": cls._text(detail.get("account")),
                "description": cls._text(detail.get("description"), "未知科目"),
                "ktosl": cls._text(detail.get("ktosl")),
                "komok": cls._text(detail.get("komok")),
                "bwmod": cls._text(detail.get("bwmod")),
                "bklas": cls._text(detail.get("bklas")),
            })
        if rows:
            return rows
        for label in scenario.get("accounts", []) or scenario.get("raw_accounts", []) or []:
            account, description = cls._parse_account_label(label)
            rows.append({
                "direction": "",
                "account": account,
                "description": description or cls._text(label),
                "ktosl": "",
                "komok": "",
                "bwmod": "",
                "bklas": "",
            })
        return rows

    @classmethod
    def _parse_account_label(cls, label):
        text = cls._text(label)
        if not text:
            return "", ""
        match = re.match(r"^(\d+)\s*\((.*)\)$", text)
        if match:
            return match.group(1), match.group(2)
        parts = text.split(maxsplit=1)
        return parts[0], parts[1].strip("() ") if len(parts) > 1 else ""

    @classmethod
    def _is_placeholder_account(cls, value):
        text = cls._text(value)
        if not text:
            return True
        return "OCR未识别" in text or "未识别对方科目" in text

    @classmethod
    def _sample_to_toe_rows(cls, sample):
        doc_num = cls._text(sample.get("DOC_NUM"))
        date = cls._text(sample.get("DATE"), "待补充")
        company_code = cls._text(sample.get("COMPANY_CODE"))
        rows = []
        for direction, key in (("借方", "DEBIT_LINES"), ("贷方", "CREDIT_LINES")):
            for line in sample.get(key) or []:
                rows.append({
                    "doc_num": doc_num,
                    "company_code": company_code,
                    "direction": direction,
                    "account": cls._text(line.get("account")),
                    "material": cls._text(line.get("matnr")),
                    "description": cls._text(line.get("description"), "未知科目"),
                    "amount": cls._num(line.get("amount")),
                    "date": date,
                    "ktosl": cls._text(line.get("ktosl") or sample.get("KTOSL")),
                    "komok": cls._text(line.get("komok") or sample.get("KOMOK")),
                })
        if rows:
            return rows
        for direction, account_key, desc_key, matnr_key in (
            ("借方", "DEBIT_ACC", "DEBIT_DESC", "DEBIT_MATNR"),
            ("贷方", "CREDIT_ACC", "CREDIT_DESC", "CREDIT_MATNR"),
        ):
            account = sample.get(account_key, "")
            if cls._is_placeholder_account(account):
                continue
            rows.append({
                "doc_num": doc_num,
                "company_code": company_code,
                "direction": direction,
                "account": cls._text(account),
                "material": cls._text(sample.get(matnr_key) or sample.get("MATNR")),
                "description": cls._text(sample.get(desc_key), "未知科目"),
                "amount": cls._num(sample.get("AMOUNT")),
                "date": date,
                "ktosl": cls._text(sample.get("KTOSL")),
                "komok": cls._text(sample.get("KOMOK")),
            })
        if rows:
            return rows
        return [{
            "doc_num": doc_num,
            "company_code": company_code,
            "direction": cls._direction_label(sample.get("SHKZG")),
            "account": cls._text(sample.get("SAKNR") or sample.get("DEBIT_ACC")),
            "material": cls._text(sample.get("MATNR") or sample.get("DEBIT_MATNR")),
            "description": cls._text(sample.get("TXT50") or sample.get("DEBIT_DESC"), "未知科目"),
            "amount": cls._num(sample.get("AMOUNT")),
            "date": date,
            "ktosl": cls._text(sample.get("KTOSL")),
            "komok": cls._text(sample.get("KOMOK")),
        }]

    @staticmethod
    def _direction_label(value):
        text = str(value or "").upper().strip()
        if text in {"H", "CR", "CREDIT", "贷", "贷方"}:
            return "贷方"
        if text in {"S", "DR", "DEBIT", "借", "借方"}:
            return "借方"
        return ""

    @classmethod
    def _group_di_results(cls, di_results):
        grouped = defaultdict(list)
        for result in di_results or []:
            grouped[cls._text(result.get("scenario"), "未指定场景")].append(result)
        return grouped

    def _fallback_sample_rows(self, scenario_name):
        samples = self._load_session_table("Samples")
        if samples.empty:
            return []
        if "SCENARIO" in samples.columns:
            scenario_mask = samples["SCENARIO"].astype(str).str.strip().eq(scenario_name)
            samples = samples[scenario_mask].copy()
        rows = []
        for _, row in samples.iterrows():
            rows.extend(self._sample_to_toe_rows(row.to_dict()))
        return rows

    @classmethod
    def _scenario_names(cls, ranked_scenarios, di_results):
        names = []
        for scenario in ranked_scenarios or []:
            name = cls._text(scenario.get("name"))
            if name and name not in names:
                names.append(name)
        for result in di_results or []:
            name = cls._text(result.get("scenario"))
            if name and name not in names:
                names.append(name)
        return names or ["未指定场景"]

    @classmethod
    def _validation_value(cls, row, *names):
        for name in names:
            if name in row:
                return cls._text(row.get(name))
        return ""

    def _write_summary(self, wb, ranked_scenarios, di_results, audit_context, styles):
        ws = wb.active
        ws.title = "审计摘要"
        ws.merge_cells("A1:H1")
        ws["A1"] = "智审 V.A.S.T. 自动化凭证审计底稿 - Executive Summary"
        ws["A1"].font = styles["title_font"]
        ws["A1"].alignment = styles["center"]

        context_rows = [
            ("被审计单位", audit_context.get("entity_name") or "待补充", "系统版本", audit_context.get("system_version") or audit_context.get("system_name") or "待补充"),
            ("审计期间", f"{audit_context.get('period_start') or '待补充'} 至 {audit_context.get('period_end') or '待补充'}", "报告定位", "SAP 自动化凭证场景识别、配置验证、异常筛查和底稿化"),
        ]
        for row_idx, values in enumerate(context_rows, start=3):
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = styles["border"]
                cell.alignment = styles["wrap"]
                if col_idx in {1, 3}:
                    cell.fill = styles["light_grey_fill"]
                    cell.font = styles["bold_font"]

        scenario_totals = sorted(
            [
                {
                    "name": self._text(item.get("name")),
                    "amount": self._scenario_total(item),
                    "company_count": len(item.get("company_values", []) or []),
                }
                for item in ranked_scenarios or []
            ],
            key=lambda item: abs(item["amount"]),
            reverse=True,
        )
        account_rows = self._account_rows(ranked_scenarios)
        top_accounts = self._top_accounts(ranked_scenarios, limit=5)
        sample_counts = defaultdict(int)
        for result in di_results or []:
            sample_counts[self._text(result.get("scenario"))] += 1
        ledger_summary = audit_context.get("full_ledger_summary") or {}

        metrics = [
            ("识别业务场景", len(ranked_scenarios or [])),
            ("样本覆盖场景", sum(1 for value in sample_counts.values() if value)),
            ("场景归集金额", sum(item["amount"] for item in scenario_totals)),
            ("命中科目记录", len(account_rows)),
            ("全量凭证金额覆盖率", f"{self._num(ledger_summary.get('amount_coverage_pct')):.2f}%" if ledger_summary else "待上传全量序时账"),
        ]
        ws["A6"] = "关键审计指标"
        ws["A6"].font = styles["section_font"]
        for idx, (label, value) in enumerate(metrics, start=1):
            label_cell = ws.cell(row=7, column=idx, value=label)
            value_cell = ws.cell(row=8, column=idx, value=value)
            label_cell.fill = styles["light_teal_fill"]
            label_cell.font = Font(bold=True, color=self.KPMG_BLUE)
            value_cell.font = Font(size=13, bold=True)
            for cell in (label_cell, value_cell):
                cell.alignment = styles["center"]
                cell.border = styles["border"]
            if isinstance(value, (int, float)):
                value_cell.number_format = "#,##0.00"

        top_scenario = scenario_totals[0] if scenario_totals else None
        top_account = top_accounts[0] if top_accounts else None
        validation_rows = audit_context.get("voucher_validation") or []
        non_pass_count = sum(
            1 for row in validation_rows
            if self._validation_value(row, "校验结论") not in {"", "通过"}
        )
        summary_lines = [
            f"1. 自动化凭证影响金额最大的场景：{top_scenario['name']}（{top_scenario['amount']:,.2f}）。" if top_scenario else "1. 暂无可量化场景金额。",
            f"2. 金额贡献最高的科目：{top_account['scenario']} - {top_account['account']} {top_account['description']}（{top_account['amount']:,.2f}）。" if top_account else "2. 暂无可量化科目金额。",
            f"3. 已形成 {len(di_results or [])} 条样本证据叙述，覆盖 {sum(1 for value in sample_counts.values() if value)} 个审计场景。",
            f"4. 凭证到 T030 配置验证中有 {non_pass_count} 条待补充/待核对事项，详见“异常_待补充清单”。",
            "5. 本底稿把 SAP 自动化凭证配置、科目主数据、物料评估信息、样本凭证与审计结论串联，用于支持财务审计团队理解自动化凭证如何影响财务报表科目。",
        ]
        if ledger_summary:
            summary_lines.append(
                f"6. 全量序时账覆盖：共 {int(self._num(ledger_summary.get('total_lines'))):,} 行，"
                f"已覆盖 {int(self._num(ledger_summary.get('covered_lines'))):,} 行，"
                f"金额覆盖率 {self._num(ledger_summary.get('amount_coverage_pct')):.2f}%。"
            )
        ws["A11"] = "审计价值摘要"
        ws["A11"].font = styles["section_font"]
        ws.merge_cells("A12:H18")
        ws["A12"] = "\n".join(summary_lines)
        ws["A12"].alignment = styles["wrap"]
        ws["A12"].fill = styles["light_grey_fill"]
        ws["A12"].border = styles["border"]

        rows = [
            [item["name"], item["amount"], item["company_count"], sample_counts.get(item["name"], 0), "优先覆盖金额重大场景" if idx == 0 else "结合金额与样本覆盖判断"]
            for idx, item in enumerate(scenario_totals[:10])
        ]
        self._write_table(ws, 21, 1, ["审计场景", "归集金额", "命中公司数", "样本数", "审计关注点"], rows, styles, "场景金额与样本覆盖")

        account_rows_for_sheet = [
            [item["scenario"], item["account"], item["description"], item["amount"]]
            for item in top_accounts
        ]
        self._write_table(ws, 35, 1, ["审计场景", "科目编码", "科目描述", "金额"], account_rows_for_sheet, styles, "重点科目贡献 Top 5")
        self._column_widths(ws, {"A": 24, "B": 18, "C": 18, "D": 18, "E": 34, "F": 24, "G": 18, "H": 24})
        ws.row_dimensions[12].height = 120

    def _write_di_sheet(self, wb, scenario_name, scenario, results, styles, used_titles):
        ws = wb.create_sheet(self._safe_sheet_title(scenario_name, used_titles))
        ws.merge_cells("D2:H2")
        ws["D2"] = "设计和执行(D&I)"
        ws["D2"].font = Font(size=14, bold=True)
        ws["D2"].alignment = styles["center"]
        ws.merge_cells("D3:H3")
        ws["D3"] = "了解流程控制活动 / Understand the process control activities"
        ws["D3"].font = styles["bold_font"]
        ws["D3"].alignment = styles["center"]

        base_rows = [
            (6, "控制 / Control", f"{scenario_name} 自动化凭证生成逻辑"),
            (8, "控制说明", self._control_description(scenario_name, scenario)),
            (20, "性质 / Nature", "自动化【AUTOMATED】"),
            (22, "类型 / Type", "预防性【PREVENTIVE】"),
        ]
        for row_idx, label, value in base_rows:
            ws.cell(row=row_idx, column=4, value=label).fill = styles["blue_fill"]
            ws.cell(row=row_idx, column=4).font = styles["header_font"]
            ws.cell(row=row_idx, column=5, value=value)
            for col in range(4, 9):
                ws.cell(row=row_idx, column=col).border = styles["border"]
                ws.cell(row=row_idx, column=col).alignment = styles["wrap"]

        prp_headers = ["流程风险点 / PRP ID", "流程风险点 / PRP(s)", "重大错报风险 / RMM", "信息 / Information", "控制活动如何应对风险"]
        prp_values = [
            f"PRP_{scenario_name}",
            f"{scenario_name} 自动生成凭证未按 SAP 配置计入正确财务科目。",
            "财务报表科目分类、发生额或截止可能存在错报。",
            "T030/SKAT/T001K/MARC/MM03、样本凭证及余额/发生额数据。",
            results[0].get("di_description") if results else "本场景已纳入 SAP 自动化凭证实质性测试分析；尚未匹配到样本凭证，需补充样本后完成 TOE。",
        ]
        for offset, header in enumerate(prp_headers):
            cell = ws.cell(row=17, column=4 + offset, value=header)
            cell.fill = styles["blue_fill"]
            cell.font = styles["header_font"]
            cell.alignment = styles["center"]
            cell.border = styles["border"]
            value_cell = ws.cell(row=18, column=4 + offset, value=prp_values[offset])
            value_cell.alignment = styles["wrap"]
            value_cell.border = styles["border"]
        ws.row_dimensions[18].height = 110

        account_rows = [
            [
                detail["direction"] or "待补充",
                detail["account"],
                detail["description"],
                detail["ktosl"],
                detail["komok"],
                detail["bwmod"],
                detail["bklas"],
                "凭证科目、借贷方向、评估分组/分类应与 T030 配置一致。",
            ]
            for detail in self._scenario_account_details(scenario)
        ]
        end_row = self._write_table(
            ws,
            25,
            4,
            ["配置借贷方", "科目编码", "科目描述", "KTOSL", "KOMOK", "评估分组", "评估类", "凭证要求"],
            account_rows,
            styles,
            "一、场景借贷科目及凭证要求",
        )

        toe_rows = []
        for result in results:
            for row in self._sample_to_toe_rows(result.get("sample_table", {})):
                toe_rows.append([
                    row["doc_num"],
                    row["company_code"],
                    row["direction"],
                    row["account"],
                    row["description"],
                    row["material"],
                    row["amount"],
                    row["date"],
                    row["ktosl"],
                    row["komok"],
                ])
        if not toe_rows:
            for row in self._fallback_sample_rows(scenario_name):
                toe_rows.append([
                    row["doc_num"],
                    row["company_code"],
                    row["direction"],
                    row["account"],
                    row["description"],
                    row["material"],
                    row["amount"],
                    row["date"],
                    row["ktosl"],
                    row["komok"],
                ])
        self._write_table(
            ws,
            end_row + 3,
            4,
            ["凭证号", "公司代码", "借贷方向", "科目编码", "科目描述", "物料号", "金额", "日期", "KTOSL", "KOMOK"],
            toe_rows,
            styles,
            "二、测试样本明细 (TOE)",
        )

        self._column_widths(ws, {
            "D": 18, "E": 18, "F": 22, "G": 18, "H": 32, "I": 18, "J": 14, "K": 14, "L": 14, "M": 14,
        })
        ws.freeze_panes = "D17"

    def _control_description(self, scenario_name, scenario):
        details = self._scenario_account_details(scenario)
        debit_accounts = [f"{d['account']} {d['description']}" for d in details if "借" in d["direction"]]
        credit_accounts = [f"{d['account']} {d['description']}" for d in details if "贷" in d["direction"]]
        debit_text = "、".join(debit_accounts[:5]) or "按 T030 配置确定的借方科目"
        credit_text = "、".join(credit_accounts[:5]) or "按 T030 配置确定的贷方科目"
        return f"系统在执行“{scenario_name}”相关业务操作时，应依据 SAP 自动过账配置自动生成会计凭证，借方计入 {debit_text}，贷方计入 {credit_text}；审计程序验证凭证科目、借贷方向、物料评估信息与配置逻辑一致。"

    def _write_config_sheet(self, wb, ranked_scenarios, styles):
        ws = wb.create_sheet("T030&SKAT 科目配置")
        rows = []
        for scenario in ranked_scenarios or []:
            scenario_name = self._text(scenario.get("name"))
            for detail in self._scenario_account_details(scenario):
                rows.append([
                    scenario_name,
                    detail["direction"],
                    detail["ktosl"],
                    detail["komok"],
                    detail["bwmod"],
                    detail["bklas"],
                    detail["account"],
                    detail["description"],
                ])
        self._write_table(
            ws,
            1,
            1,
            ["审计场景", "配置借贷方", "KTOSL", "KOMOK", "评估分组(BWMOD)", "评估类(BKLAS)", "科目编码", "SKAT科目描述"],
            rows,
            styles,
            "T030 与 SKAT 科目配置证据",
        )
        self._column_widths(ws, {"A": 20, "B": 12, "C": 12, "D": 14, "E": 18, "F": 16, "G": 18, "H": 36})
        ws.freeze_panes = "A3"

    def _write_valuation_sheet(self, wb, audit_context, styles):
        ws = wb.create_sheet("T001K_MARC_MM03 评估信息")
        rows = []
        for row in audit_context.get("voucher_validation") or []:
            rows.append([
                self._validation_value(row, "凭证号"),
                self._validation_value(row, "审计场景"),
                self._validation_value(row, "公司代码"),
                self._validation_value(row, "T001K评估分组"),
                self._validation_value(row, "物料号"),
                self._validation_value(row, "物料主数据来源"),
                self._validation_value(row, "MARC/MM03工厂", "MM03工厂"),
                self._validation_value(row, "MARC/MM03评估分类", "MM03评估分类"),
                self._validation_value(row, "科目编码"),
                self._validation_value(row, "借贷方向"),
                self._validation_value(row, "T030期望科目"),
                self._validation_value(row, "校验结论"),
                self._validation_value(row, "校验说明"),
            ])
        if not rows:
            samples = self._load_session_table("Samples")
            for _, sample in samples.head(200).iterrows():
                rows.append([
                    self._text(sample.get("DOC_NUM")),
                    self._text(sample.get("SCENARIO")),
                    self._text(sample.get("COMPANY_CODE")),
                    "",
                    self._text(sample.get("MATNR")),
                    "",
                    "",
                    "",
                    self._text(sample.get("SAKNR")),
                    self._direction_label(sample.get("SHKZG")),
                    "",
                    "待补充",
                    "未生成凭证到 T030 配置验证结果。",
                ])
        self._write_table(
            ws,
            1,
            1,
            ["凭证号", "审计场景", "公司代码", "T001K评估分组", "物料号", "物料主数据来源", "MARC/MM03工厂", "MARC/MM03评估分类", "科目编码", "借贷方向", "T030期望科目", "校验结论", "校验说明"],
            rows,
            styles,
            "T001K / MARC / MM03 / T030 评估链路核对",
        )
        self._column_widths(ws, {"A": 18, "B": 18, "C": 14, "D": 18, "E": 18, "F": 16, "G": 16, "H": 18, "I": 16, "J": 12, "K": 30, "L": 12, "M": 46})
        ws.freeze_panes = "A3"

    def _write_sample_detail_sheet(self, wb, di_results, styles):
        ws = wb.create_sheet("样本凭证明细")
        rows = []
        for result in di_results or []:
            scenario = self._text(result.get("scenario"))
            for row in self._sample_to_toe_rows(result.get("sample_table", {})):
                rows.append([
                    scenario,
                    row["doc_num"],
                    row["company_code"],
                    row["direction"],
                    row["account"],
                    row["description"],
                    row["material"],
                    row["amount"],
                    row["date"],
                    row["ktosl"],
                    row["komok"],
                    self._text(result.get("di_description")),
                ])
        if not rows:
            samples = self._load_session_table("Samples")
            for _, sample in samples.iterrows():
                rows.append([
                    self._text(sample.get("SCENARIO")),
                    self._text(sample.get("DOC_NUM")),
                    self._text(sample.get("COMPANY_CODE")),
                    self._direction_label(sample.get("SHKZG")),
                    self._text(sample.get("SAKNR")),
                    self._text(sample.get("TXT50"), "未知科目"),
                    self._text(sample.get("MATNR")),
                    self._num(sample.get("AMOUNT")),
                    self._text(sample.get("DATE")),
                    self._text(sample.get("KTOSL")),
                    self._text(sample.get("KOMOK")),
                    "本场景未匹配到样本凭证" if samples.empty else "",
                ])
        self._write_table(
            ws,
            1,
            1,
            ["审计场景", "凭证号", "公司代码", "借贷方向", "科目编码", "科目描述", "物料号", "金额", "日期", "KTOSL", "KOMOK", "TOD/TOE 描述"],
            rows,
            styles,
            "Step 3 真实样本凭证明细",
        )
        self._column_widths(ws, {"A": 18, "B": 18, "C": 14, "D": 12, "E": 16, "F": 32, "G": 18, "H": 16, "I": 14, "J": 12, "K": 12, "L": 70})
        ws.freeze_panes = "A3"

    def _write_information_sheet(self, wb, audit_context, styles):
        ws = wb.create_sheet("Information")
        source_rows = []
        source_specs = [
            ("T030", "自动过账配置表", "识别自动化凭证规则及科目确定逻辑"),
            ("SKAT", "科目主数据", "补充科目描述"),
            ("T001K", "公司代码/评估分组", "凭证公司代码定位评估分组"),
            ("MARC", "物料主数据", "凭证物料号定位工厂与评估分类；MM03 截图仅作为补充证据"),
            ("Samples", "样本凭证清单", "生成 TOE 明细和凭证验证"),
            ("TrialBalance", "科目余额/发生额表", "量化场景金额和科目占比"),
        ]
        for table_name, source_type, purpose in source_specs:
            df = self._load_session_table(table_name)
            source_rows.append([
                table_name,
                source_type,
                len(df.index) if not df.empty else 0,
                "已获取" if not df.empty else "未上传/未识别",
                purpose,
                "电子数据；由项目资料包或用户上传文件解析生成。",
            ])
        if audit_context.get("full_ledger_summary"):
            source_rows.append([
                "Ledger",
                "全量序时账/凭证明细",
                int(self._num(audit_context["full_ledger_summary"].get("total_lines"))),
                "已获取",
                "执行全量自动化凭证场景归类和异常筛选",
                "电子数据；用于扩展实质性测试覆盖。",
            ])
        self._write_table(
            ws,
            1,
            1,
            ["数据对象", "数据类型", "记录数", "获取状态", "审计用途", "信息来源/可靠性程序"],
            source_rows,
            styles,
            "Related Information to be tested",
        )
        self._column_widths(ws, {"A": 16, "B": 22, "C": 12, "D": 14, "E": 42, "F": 54})
        ws.freeze_panes = "A3"

    def _write_exception_sheet(self, wb, audit_context, styles):
        ws = wb.create_sheet("异常_待补充清单")
        rows = []
        for row in audit_context.get("voucher_validation") or []:
            conclusion = self._validation_value(row, "校验结论")
            if conclusion and conclusion != "通过":
                rows.append([
                    "凭证配置验证",
                    self._validation_value(row, "凭证号"),
                    self._validation_value(row, "审计场景"),
                    self._validation_value(row, "公司代码"),
                    self._validation_value(row, "科目编码"),
                    conclusion,
                    self._validation_value(row, "校验说明"),
                ])
        for row in audit_context.get("full_ledger_exceptions") or []:
            rows.append([
                "全量序时账异常",
                self._validation_value(row, "凭证号", "DOC_NUM", "BELNR"),
                self._validation_value(row, "审计场景", "SCENARIO"),
                self._validation_value(row, "公司代码", "COMPANY_CODE", "BUKRS"),
                self._validation_value(row, "科目编码", "SAKNR", "总账科目"),
                self._validation_value(row, "状态", "结论") or "待核查",
                self._validation_value(row, "说明", "原因") or "未纳入自动化凭证测试覆盖或配置验证未通过。",
            ])
        self._write_table(
            ws,
            1,
            1,
            ["来源", "凭证号", "审计场景", "公司代码", "科目编码", "结论", "说明"],
            rows,
            styles,
            "异常 / 待补充事项清单",
        )
        self._column_widths(ws, {"A": 18, "B": 18, "C": 18, "D": 14, "E": 16, "F": 14, "G": 60})
        ws.freeze_panes = "A3"

    def generate(self, ranked_scenarios, di_results, audit_context=None):
        audit_context = audit_context or {}
        wb = Workbook()
        styles = self._styles()
        self._write_summary(wb, ranked_scenarios, di_results, audit_context, styles)

        grouped_di = self._group_di_results(di_results)
        scenario_lookup = {self._text(item.get("name")): item for item in ranked_scenarios or []}
        used_titles = {wb.active.title}
        for scenario_name in self._scenario_names(ranked_scenarios, di_results):
            self._write_di_sheet(
                wb,
                scenario_name,
                scenario_lookup.get(scenario_name, {"name": scenario_name}),
                grouped_di.get(scenario_name, []),
                styles,
                used_titles,
            )

        self._write_config_sheet(wb, ranked_scenarios, styles)
        self._write_valuation_sheet(wb, audit_context, styles)
        self._write_sample_detail_sheet(wb, di_results, styles)
        self._write_information_sheet(wb, audit_context, styles)
        self._write_exception_sheet(wb, audit_context, styles)

        wb.save(self.output_path)
        return self.output_path
