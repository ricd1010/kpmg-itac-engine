from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from report_generator import ReportGenerator


class GeneralAuditReportGenerator(ReportGenerator):
    @staticmethod
    def _num(value):
        try:
            return float(value or 0)
        except Exception:
            return 0.0

    @classmethod
    def _scenario_total(cls, scenario):
        if "total_value" in scenario:
            return cls._num(scenario.get("total_value"))
        total = 0.0
        for company in scenario.get("company_values", []) or []:
            total += cls._num(company.get("total_value"))
            if not company.get("total_value"):
                for account in company.get("account_values", []) or []:
                    total += cls._num(account.get("amount")) or cls._num(account.get("value"))
        return total

    @classmethod
    def _account_rows(cls, ranked_scenarios):
        rows = []
        for scenario in ranked_scenarios or []:
            scenario_name = str(scenario.get("name", "") or "")
            for company in scenario.get("company_values", []) or []:
                company_code = str(company.get("company_code", "未指定公司") or "未指定公司")
                for account in company.get("account_values", []) or []:
                    total = (
                        cls._num(account.get("amount"))
                        or cls._num(account.get("total_value"))
                        or cls._num(account.get("value"))
                    )
                    debit = cls._num(account.get("debit_value"))
                    credit = cls._num(account.get("credit_value"))
                    if not (total or debit or credit):
                        continue
                    rows.append({
                        "scenario": scenario_name,
                        "company": company_code,
                        "account": str(account.get("account", "") or ""),
                        "description": str(account.get("description", "未知科目") or "未知科目"),
                        "amount": total or debit + credit,
                    })
        return rows

    @classmethod
    def _top_accounts(cls, ranked_scenarios, limit=5):
        totals = defaultdict(float)
        descriptions = {}
        for row in cls._account_rows(ranked_scenarios):
            key = (row["scenario"], row["account"])
            totals[key] += cls._num(row["amount"])
            descriptions[key] = row["description"]
        ranked = sorted(totals.items(), key=lambda item: abs(item[1]), reverse=True)
        return [
            {
                "scenario": scenario,
                "account": account,
                "description": descriptions.get((scenario, account), ""),
                "amount": amount,
            }
            for (scenario, account), amount in ranked[:limit]
        ]

    @classmethod
    def _company_concentration_summary(cls, ranked_scenarios, limit=5):
        totals = defaultdict(float)
        for row in cls._account_rows(ranked_scenarios):
            totals[row["company"]] += cls._num(row["amount"])
        return sorted(totals.items(), key=lambda item: (-abs(item[1]), item[0]))[:limit]

    @staticmethod
    def _sample_counts(di_results):
        counts = defaultdict(int)
        for item in di_results or []:
            counts[str(item.get("scenario", "") or "")] += 1
        return counts

    def _write_voucher_validation_sheet(self, wb, rows, styles):
        if not rows:
            return
        ws = wb.create_sheet("凭证配置验证")
        blue_fill = styles["blue_fill"]
        header_font = styles["header_font"]
        border = styles["border"]
        align_wrap = styles["align_wrap"]
        section_font = styles["section_font"]

        headers = [
            "凭证号", "审计场景", "公司代码", "物料号", "科目编码", "借贷方向",
            "T001K评估分组", "MM03工厂", "MM03评估分类", "T030期望科目", "校验结论", "校验说明",
        ]
        ws["A1"] = "凭证到 T030 配置验证"
        ws["A1"].font = section_font
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = blue_fill
            cell.font = header_font
            cell.alignment = align_wrap
            cell.border = border
        for row_idx, row in enumerate(rows, start=4):
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col, value=row.get(header, ""))
                cell.alignment = align_wrap
                cell.border = border
        widths = {
            "A": 18, "B": 16, "C": 12, "D": 18, "E": 16, "F": 12,
            "G": 16, "H": 12, "I": 14, "J": 28, "K": 12, "L": 46,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    def generate(self, ranked_scenarios, di_results, audit_context=None):
        path = super().generate(ranked_scenarios, di_results, audit_context)
        audit_context = audit_context or {}

        wb = load_workbook(path)
        if wb.worksheets:
            wb.worksheets[0].title = "技术审计总览"
        ws = wb.create_sheet("Executive Summary", 0)

        kpmg_blue = "00338D"
        teal_fill = PatternFill(start_color="EAF7F7", end_color="EAF7F7", fill_type="solid")
        blue_fill = PatternFill(start_color=kpmg_blue, end_color=kpmg_blue, fill_type="solid")
        light_fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        title_font = Font(size=18, bold=True, color=kpmg_blue)
        section_font = Font(size=12, bold=True, color=kpmg_blue)
        align_wrap = Alignment(wrap_text=True, vertical="top")
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="thin", color="D8DDE6"),
            right=Side(style="thin", color="D8DDE6"),
            top=Side(style="thin", color="D8DDE6"),
            bottom=Side(style="thin", color="D8DDE6"),
        )

        ws.merge_cells("A1:F1")
        ws["A1"] = "TSDA 测试范围框定辅助驾驶舱 Executive Summary"
        ws["A1"].font = title_font
        ws["A1"].alignment = align_center

        ws["A3"] = "被审计单位"
        ws["B3"] = audit_context.get("entity_name", "")
        ws["D3"] = "系统版本"
        ws["E3"] = audit_context.get("system_version") or audit_context.get("system_name", "")
        ws["A4"] = "审计期间"
        ws["B4"] = f"{audit_context.get('period_start', '')} 至 {audit_context.get('period_end', '')}"
        ws["D4"] = "报告定位"
        ws["E4"] = "面向财务审计与 IT 审计的自动过账、科目归集和样本证据摘要"

        for cell in ["A3", "D3", "A4", "D4"]:
            ws[cell].font = Font(bold=True)
            ws[cell].fill = light_fill

        scenario_totals = sorted(
            [
                {
                    "name": str(item.get("name", "") or ""),
                    "amount": self._scenario_total(item),
                    "company_count": len(item.get("company_values", []) or []),
                }
                for item in ranked_scenarios or []
            ],
            key=lambda item: abs(item["amount"]),
            reverse=True,
        )
        account_rows = self._account_rows(ranked_scenarios)
        top_accounts = self._top_accounts(ranked_scenarios)
        concentrated_companies = self._company_concentration_summary(ranked_scenarios)
        sample_counts = self._sample_counts(di_results)
        covered_scenarios = sum(1 for count in sample_counts.values() if count)

        metrics = [
            ("识别业务场景", len(ranked_scenarios or [])),
            ("样本覆盖场景", covered_scenarios),
            ("归集金额", f"{sum(item['amount'] for item in scenario_totals):,.2f}"),
            ("命中科目记录", len(account_rows)),
            ("重点公司数", len(concentrated_companies)),
        ]
        ws["A6"] = "关键审计指标"
        ws["A6"].font = section_font
        for idx, (label, value) in enumerate(metrics, start=1):
            col = idx
            ws.cell(row=7, column=col, value=label)
            ws.cell(row=8, column=col, value=value)
            ws.cell(row=7, column=col).fill = teal_fill
            ws.cell(row=7, column=col).font = Font(bold=True, color=kpmg_blue)
            ws.cell(row=8, column=col).font = Font(size=13, bold=True)
            ws.cell(row=7, column=col).alignment = align_center
            ws.cell(row=8, column=col).alignment = align_center
            ws.cell(row=7, column=col).border = border
            ws.cell(row=8, column=col).border = border

        top_scenario_text = (
            f"{scenario_totals[0]['name']}（{scenario_totals[0]['amount']:,.2f}）"
            if scenario_totals else "暂无金额数据"
        )
        top_account_text = (
            f"{top_accounts[0]['scenario']} - {top_accounts[0]['account']} {top_accounts[0]['description']}（{top_accounts[0]['amount']:,.2f}）"
            if top_accounts else "暂无科目金额数据"
        )
        concentrated_company_text = "、".join(
            f"{company}（{amount:,.2f}）"
            for company, amount in concentrated_companies[:3]
        ) or "暂无公司维度金额数据"
        sample_focus_text = "；".join(
            f"{row['scenario']} - {row['account']} {row['description']}"
            for row in top_accounts[:3]
        ) or "待上传样本后补充"

        ws["A11"] = "审计价值摘要"
        ws["A11"].font = section_font
        summary_lines = [
            f"1. SAP 自动分录金额影响最大的业务场景：{top_scenario_text}。",
            f"2. 金额贡献最高的财务科目：{top_account_text}。",
            f"3. 自动分录金额较集中的公司：{concentrated_company_text}，建议结合样本覆盖情况安排穿行测试。",
            f"4. 建议优先抽样覆盖对象：{sample_focus_text}。",
            "5. 本底稿将自动过账配置、科目主数据、余额/发生额和样本凭证串联，用于支持审计团队理解自动分录如何影响财务报表科目。",
        ]
        ws.merge_cells("A12:F17")
        ws["A12"] = "\n".join(summary_lines)
        ws["A12"].alignment = align_wrap
        ws["A12"].fill = light_fill
        ws["A12"].border = border

        ws["A20"] = "场景金额与样本覆盖"
        ws["A20"].font = section_font
        headers = ["审计场景", "归集金额", "命中公司数", "样本数", "审计关注点"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=21, column=col, value=header)
            cell.fill = blue_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = border

        for row_idx, item in enumerate(scenario_totals[:10], start=22):
            sample_count = sample_counts.get(item["name"], 0)
            focus = "优先覆盖金额重大场景" if row_idx == 22 else "结合金额与样本覆盖情况判断"
            values = [item["name"], item["amount"], item["company_count"], sample_count, focus]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = align_wrap
                if col == 2:
                    cell.number_format = '#,##0.00'

        ws["A35"] = "重点科目贡献 Top 5"
        ws["A35"].font = section_font
        account_headers = ["审计场景", "科目编码", "科目描述", "金额"]
        for col, header in enumerate(account_headers, start=1):
            cell = ws.cell(row=36, column=col, value=header)
            cell.fill = blue_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = border
        for row_idx, item in enumerate(top_accounts, start=37):
            values = [item["scenario"], item["account"], item["description"], item["amount"]]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = align_wrap
                if col == 4:
                    cell.number_format = '#,##0.00'

        widths = {"A": 24, "B": 20, "C": 28, "D": 18, "E": 34, "F": 28}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        for row in range(1, 45):
            ws.row_dimensions[row].height = 22
        ws.row_dimensions[12].height = 120

        self._write_voucher_validation_sheet(
            wb,
            audit_context.get("voucher_validation") or [],
            {
                "blue_fill": blue_fill,
                "header_font": header_font,
                "border": border,
                "align_wrap": align_wrap,
                "section_font": section_font,
            },
        )

        wb.save(path)
        return path
