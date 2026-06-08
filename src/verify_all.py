import os
from core1_main import Core1Orchestrator
from core2_main import Core2Orchestrator
from report_generator import ReportGenerator
import openpyxl

def verify_all():
    data_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data")
    
    print(">>> 正在启动验证流程...")
    
    # Audit Context
    audit_context = {
        "entity_name": "测试有限公司",
        "system_name": "SAP ERP",
        "period_start": "2025-01-01",
        "period_end": "2025-12-31"
    }

    # 1. Core 1
    print(">>> 验证步骤 1: 场景识别...")
    c1 = Core1Orchestrator(data_dir)
    ranked_scenarios = c1.run()
    if not ranked_scenarios:
        raise Exception("Core 1 未识别到场景")
    print(f"    Core 1 识别到 {len(ranked_scenarios)} 个场景")

    # 2. Core 2
    print(">>> 验证步骤 2: AI 描述生成...")
    c2 = Core2Orchestrator(data_dir)
    di_results = c2.generate_di_descriptions(ranked_scenarios, audit_context)
    if not di_results:
        raise Exception("Core 2 未生成描述")
    print(f"    Core 2 生成了 {len(di_results)} 个描述")

    # 3. Core 3
    print(">>> 验证步骤 3: Excel 底稿生成...")
    generator = ReportGenerator(data_dir)
    report_path = generator.generate(ranked_scenarios, di_results, audit_context)
    
    if not os.path.exists(report_path):
        raise Exception("Excel 文件未生成")
    
    # 4. Verify Excel Integrity
    print(">>> 验证步骤 4: 检查 Excel 文件完整性...")
    try:
        wb = openpyxl.load_workbook(report_path)
        print(f"    成功加载 Excel，包含分页: {wb.sheetnames}")
        if "审计总览" not in wb.sheetnames:
            raise Exception("缺失 '审计总览' 分页")
        if "方法论与较佳实践 (WGLL)" not in wb.sheetnames:
            raise Exception("缺失 'WGLL' 分页")
        print("    Excel 文件结构校验通过")
    except Exception as e:
        raise Exception(f"Excel 文件损坏或无法打开: {str(e)}")

    print("\n✅ 所有功能测试通过！最终底稿 Excel 格式校验成功。")

if __name__ == "__main__":
    verify_all()
