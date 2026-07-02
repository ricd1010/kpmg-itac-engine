import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "src"))

from report_generator import ReportGenerator


def _cell_values(ws):
    values = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                values.append(str(cell.value))
    return values


def test_report_generator_creates_standard_working_paper_sheets(tmp_path):
    generator = ReportGenerator(tmp_path)

    path = generator.generate(
        ranked_scenarios=[
            {
                "name": "采购收货",
                "total_value": 1000,
                "account_details": [
                    {
                        "direction": "借方",
                        "account": "1405020000",
                        "description": "库存商品-自制成品",
                        "ktosl": "BSX",
                        "komok": "",
                        "bwmod": "4020",
                        "bklas": "3000",
                    },
                    {
                        "direction": "贷方",
                        "account": "2202040000",
                        "description": "应付账款-GR/IR",
                        "ktosl": "WRX",
                        "komok": "",
                        "bwmod": "4020",
                        "bklas": "3000",
                    },
                ],
            }
        ],
        di_results=[],
        audit_context={"entity_name": "测试公司", "system_version": "SAP ECC"},
    )

    wb = load_workbook(path)

    for sheet in [
        "审计摘要",
        "采购收货",
        "T030&SKAT 科目配置",
        "T001K_MARC_MM03 评估信息",
        "样本凭证明细",
        "Information",
        "异常_待补充清单",
    ]:
        assert sheet in wb.sheetnames

    assert wb["审计摘要"]["A1"].value == "智审 V.A.S.T. 自动化凭证审计底稿 - Executive Summary"
    assert wb["采购收货"]["D2"].value == "设计和执行(D&I)"
    assert wb["采购收货"]["D6"].value == "控制 / Control"


def test_report_generator_expands_pair_samples_to_voucher_lines(tmp_path):
    generator = ReportGenerator(tmp_path)

    path = generator.generate(
        ranked_scenarios=[{"name": "销售成本结转", "total_value": 0}],
        di_results=[
            {
                "scenario": "销售成本结转",
                "di_description": "description",
                "sample_table": {
                    "DOC_NUM": "8174291462",
                    "COMPANY_CODE": "4000",
                    "DATE": "2026-01-08",
                    "DEBIT_ACC": "6401000000",
                    "DEBIT_MATNR": "MAT-A",
                    "DEBIT_DESC": "主营业务成本",
                    "CREDIT_ACC": "1405010000",
                    "CREDIT_MATNR": "MAT-A",
                    "CREDIT_DESC": "库存商品-产成品",
                    "AMOUNT": 13.21,
                },
            }
        ],
    )

    ws = load_workbook(path)["样本凭证明细"]
    headers = [ws.cell(row=2, column=col).value for col in range(1, 13)]
    assert headers == [
        "审计场景",
        "凭证号",
        "公司代码",
        "借贷方向",
        "科目编码",
        "科目描述",
        "物料号",
        "金额",
        "日期",
        "KTOSL",
        "KOMOK",
        "TOD/TOE 描述",
    ]
    assert [ws.cell(row=row, column=4).value for row in range(3, 5)] == ["借方", "贷方"]
    assert [ws.cell(row=row, column=5).value for row in range(3, 5)] == ["6401000000", "1405010000"]
    assert [ws.cell(row=row, column=7).value for row in range(3, 5)] == ["MAT-A", "MAT-A"]
    assert [ws.cell(row=row, column=8).value for row in range(3, 5)] == [13.21, 13.21]


def test_report_generator_uses_multiline_sample_details(tmp_path):
    generator = ReportGenerator(tmp_path)

    path = generator.generate(
        ranked_scenarios=[{"name": "采购收货", "total_value": 0}],
        di_results=[
            {
                "scenario": "采购收货",
                "di_description": "description",
                "sample_table": {
                    "DOC_NUM": "6000004976",
                    "COMPANY_CODE": "4020",
                    "DATE": "2025-07-28",
                    "DEBIT_ACC": "1405020000; 1405050100",
                    "DEBIT_DESC": "库存商品-自制成品; 库存商品-自制成品差异-采购差异",
                    "CREDIT_ACC": "2202040000",
                    "CREDIT_DESC": "应付账款-GR/IR",
                    "AMOUNT": 598470.6,
                    "DEBIT_LINES": [
                        {"account": "1405020000", "matnr": "TX5F6609-0000", "description": "库存商品-自制成品", "amount": 528470.34},
                        {"account": "1405050100", "matnr": "TX5F6609-0000", "description": "库存商品-自制成品差异-采购差异", "amount": 70000.26},
                    ],
                    "CREDIT_LINES": [
                        {"account": "2202040000", "matnr": "TX5F6609-0000", "description": "应付账款-GR/IR", "amount": 598470.6},
                    ],
                },
            },
        ],
    )

    ws = load_workbook(path)["样本凭证明细"]

    assert [ws.cell(row=row, column=4).value for row in range(3, 6)] == ["借方", "借方", "贷方"]
    assert [ws.cell(row=row, column=5).value for row in range(3, 6)] == [
        "1405020000",
        "1405050100",
        "2202040000",
    ]
    assert [ws.cell(row=row, column=7).value for row in range(3, 6)] == ["TX5F6609-0000", "TX5F6609-0000", "TX5F6609-0000"]
    assert [ws.cell(row=row, column=8).value for row in range(3, 6)] == [528470.34, 70000.26, 598470.6]


def test_report_generator_writes_information_and_exceptions_without_template_project_data(tmp_path):
    pd.DataFrame({"KTOSL": ["BSX"], "KONTS": ["1405020000"]}).to_csv(tmp_path / "T030.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame({"SAKNR": ["1405020000"], "TXT50": ["库存商品-自制成品"]}).to_csv(tmp_path / "SKAT.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame({"BUKRS": ["4020"], "BWMOD": ["4020"]}).to_csv(tmp_path / "T001K.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame({"DOC_NUM": ["6000004976"], "SCENARIO": ["采购收货"], "SAKNR": ["1405020000"]}).to_csv(tmp_path / "Samples.csv", index=False, encoding="utf-8-sig")

    generator = ReportGenerator(tmp_path)
    path = generator.generate(
        ranked_scenarios=[{"name": "采购收货", "total_value": 0}],
        di_results=[],
        audit_context={
            "voucher_validation": [
                {
                    "凭证号": "6000004976",
                    "审计场景": "采购收货",
                    "公司代码": "",
                    "科目编码": "1405020000",
                    "校验结论": "待补充",
                    "校验说明": "缺少凭证公司代码，无法定位 T001K 评估分组。",
                }
            ]
        },
    )

    wb = load_workbook(path)
    all_values = "\n".join(value for ws in wb.worksheets for value in _cell_values(ws))

    assert "缺少凭证公司代码" in all_values
    assert wb["Information"].cell(row=3, column=1).value == "T030"
    assert "曾唯" not in all_values
    assert "李忠斌" not in all_values
    assert "19983301224" not in all_values
    assert "17208270757" not in all_values


def test_di_sheet_uses_samples_csv_when_ai_description_is_empty(tmp_path):
    pd.DataFrame({
        "SCENARIO": ["采购收货"],
        "DOC_NUM": ["4700052972"],
        "COMPANY_CODE": ["4020"],
        "SAKNR": ["1403010000"],
        "TXT50": ["原材料"],
        "MATNR": ["10000000"],
        "AMOUNT": ["528470.34"],
        "SHKZG": ["S"],
        "DATE": ["2026-06-24"],
    }).to_csv(tmp_path / "Samples.csv", index=False, encoding="utf-8-sig")

    path = ReportGenerator(tmp_path).generate(
        ranked_scenarios=[{"name": "采购收货", "total_value": 0}],
        di_results=[],
        audit_context={},
    )

    ws = load_workbook(path)["采购收货"]
    values = _cell_values(ws)

    assert "4700052972" in values
    assert "1403010000" in values
    assert "原材料" in values
